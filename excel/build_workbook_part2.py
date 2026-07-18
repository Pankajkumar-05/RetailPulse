"""
Adds two sheets to the existing workbook:
  4. Reorder Point Calculator (Store 1 snapshot, formula-driven)
  5. Executive Summary (cross-sheet formula links, formatted for leadership)
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
import os

BASE = os.path.dirname(__file__)
REORDER_CSV = os.path.join(BASE, "..", "data", "processed", "excel_reorder_source.csv")
WB_PATH = os.path.join(BASE, "RetailPulse_Validation_Report.xlsx")

FONT_NAME = "Arial"
HEADER_FILL = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=16, color="2E86AB")
SUBTITLE_FONT = Font(name=FONT_NAME, italic=True, size=10, color="666666")
LABEL_FONT = Font(name=FONT_NAME, bold=True, size=11)
NORMAL_FONT = Font(name=FONT_NAME, size=10)
THIN_BORDER = Border(*[Side(style="thin", color="CCCCCC")] * 4)
INPUT_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
KPI_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

wb = openpyxl.load_workbook(WB_PATH)
df_reorder = pd.read_csv(REORDER_CSV)

# =================================================================
# SHEET 4: Reorder Point Calculator
# =================================================================
ws_ro = wb.create_sheet("Reorder Calculator")
ws_ro["A1"] = "RetailPulse — Reorder Point Calculator (Store #001 Snapshot)"
ws_ro["A1"].font = TITLE_FONT
ws_ro["A2"] = "Mirrors the SQL reorder-quantity logic (sql/analysis_queries.sql, Q14) as live Excel formulas."
ws_ro["A2"].font = SUBTITLE_FONT
ws_ro["A3"] = "Yellow cell = editable safety stock assumption. Formulas recalculate automatically."
ws_ro["A3"].font = Font(name=FONT_NAME, italic=True, size=9, color="AA6600")

# Safety stock % assumption - editable input
ws_ro["A5"] = "Safety Stock Buffer %:"
ws_ro["A5"].font = LABEL_FONT
ws_ro["B5"] = 0.2
ws_ro["B5"].number_format = "0%"
ws_ro["B5"].fill = INPUT_FILL
ws_ro["B5"].font = NORMAL_FONT

ro_headers = ["Product Name", "Category", "Stock on Hand", "Reorder Level", "Supplier",
              "Lead Time (Days)", "Avg Daily Demand (est.)", "Suggested Reorder Qty", "Action"]
ws_ro.append([])  # row 6
ws_ro.append(ro_headers)  # row 7
for cell in ws_ro[7]:
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", wrap_text=True)

start_row = 8
for i, row in df_reorder.iterrows():
    r = start_row + i
    ws_ro[f"A{r}"] = row["product_name"]
    ws_ro[f"B{r}"] = row["category"]
    ws_ro[f"C{r}"] = int(row["stock_on_hand"])
    ws_ro[f"D{r}"] = int(row["reorder_level"])
    ws_ro[f"E{r}"] = row["supplier_name"]
    ws_ro[f"F{r}"] = int(row["lead_time_days"])
    # Avg daily demand estimated from reorder_level (which was set as ~7 days demand in source data)
    ws_ro[f"G{r}"] = f"=D{r}/7"
    # Suggested reorder qty = (avg daily demand * lead time) + safety stock - current stock, floored at 0
    ws_ro[f"H{r}"] = f"=MAX(0,ROUND(G{r}*F{r}+(D{r}*$B$5)-C{r},0))"
    ws_ro[f"I{r}"] = f'=IF(C{r}<D{r},"Reorder Now","OK")'
    for col in "ABCDEFGHI":
        ws_ro[f"{col}{r}"].font = NORMAL_FONT
        ws_ro[f"{col}{r}"].border = THIN_BORDER

col_widths = {"A": 20, "B": 16, "C": 14, "D": 14, "E": 18, "F": 14, "G": 20, "H": 20, "I": 14}
for col, width in col_widths.items():
    ws_ro.column_dimensions[col].width = width

last_row = start_row + len(df_reorder) - 1
# Highlight "Reorder Now" rows
ws_ro.conditional_formatting.add(
    f"I{start_row}:I{last_row}",
    CellIsRule(operator="equal", formula=['"Reorder Now"'], fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"))
)
ws_ro.freeze_panes = "A8"

# =================================================================
# SHEET 5: Executive Summary
# =================================================================
ws_exec = wb.create_sheet("Executive Summary", 0)  # insert as first sheet
ws_exec.sheet_view.showGridLines = False

ws_exec["B2"] = "RetailPulse — Executive Summary"
ws_exec["B2"].font = Font(name=FONT_NAME, bold=True, size=20, color="2E86AB")
ws_exec["B3"] = "Q4 2024 (October – December) — Multi-Region Retail Performance"
ws_exec["B3"].font = SUBTITLE_FONT

kpi_labels = [
    ("Total Revenue", "='Regional Summary'!B10", '#,##0'),
    ("Total Gross Profit", "='Regional Summary'!D10", '#,##0'),
    ("Overall Margin %", "='Regional Summary'!E10", '0.0%'),
    ("Total Units Sold", "='Regional Summary'!F10", '#,##0'),
    ("Total Transactions", "='Regional Summary'!G10", '#,##0'),
]

row = 5
for label, formula, fmt in kpi_labels:
    ws_exec[f"B{row}"] = label
    ws_exec[f"B{row}"].font = LABEL_FONT
    ws_exec[f"C{row}"] = formula
    ws_exec[f"C{row}"].font = Font(name=FONT_NAME, bold=True, size=14, color="2E86AB")
    ws_exec[f"C{row}"].number_format = fmt
    ws_exec[f"B{row}"].fill = KPI_FILL
    ws_exec[f"C{row}"].fill = KPI_FILL
    row += 2

ws_exec[f"B{row+1}"] = "Top Region by Revenue:"
ws_exec[f"B{row+1}"].font = LABEL_FONT
ws_exec[f"C{row+1}"] = "=INDEX('Regional Summary'!A5:A9,MATCH(MAX('Regional Summary'!B5:B9),'Regional Summary'!B5:B9,0))"
ws_exec[f"C{row+1}"].font = Font(name=FONT_NAME, bold=True, size=12)

row += 3
ws_exec[f"B{row+1}"] = "Top Category by Revenue:"
ws_exec[f"B{row+1}"].font = LABEL_FONT
ws_exec[f"C{row+1}"] = "=INDEX('Category Profitability'!A5:A11,MATCH(MAX('Category Profitability'!B5:B11),'Category Profitability'!B5:B11,0))"
ws_exec[f"C{row+1}"].font = Font(name=FONT_NAME, bold=True, size=12)

row += 3
ws_exec[f"B{row+1}"] = "Products Needing Reorder (Store #001 sample):"
ws_exec[f"B{row+1}"].font = LABEL_FONT
ws_exec[f"C{row+1}"] = f"=COUNTIF('Reorder Calculator'!I8:I{last_row},\"Reorder Now\")"
ws_exec[f"C{row+1}"].font = Font(name=FONT_NAME, bold=True, size=12)

row += 4
ws_exec[f"B{row}"] = "Key Insight (from Python EDA):"
ws_exec[f"B{row}"].font = LABEL_FONT
ws_exec[f"B{row+1}"] = "Correlation between discount % and units sold is ~0.001 (near zero) —"
ws_exec[f"B{row+1}"].font = NORMAL_FONT
ws_exec[f"B{row+2}"] = "discounting is not meaningfully driving incremental volume. Recommend"
ws_exec[f"B{row+2}"].font = NORMAL_FONT
ws_exec[f"B{row+3}"] = "reassessing promo spend allocation. (See docs/eda_summary.txt)"
ws_exec[f"B{row+3}"].font = NORMAL_FONT

ws_exec.column_dimensions["A"].width = 3
ws_exec.column_dimensions["B"].width = 32
ws_exec.column_dimensions["C"].width = 22

wb.save(WB_PATH)
print(f"Saved full workbook (5 sheets) to {WB_PATH}")
