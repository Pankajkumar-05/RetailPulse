"""
Builds RetailPulse_Validation_Report.xlsx with:
  1. Raw Data sheet (Q4 2024 slice, 1,155 rows) - source of truth
  2. Regional Summary (SUMIFS formulas cross-checking SQL/Power BI totals)
  3. Category Profitability (SUMIFS + calculated margin)
  4. Reorder Point Calculator (INDEX/MATCH + formula-driven reorder logic)
  5. Executive Summary (formula-linked KPI cells, formatted for leadership)

All totals use live Excel formulas (SUMIFS/INDEX/MATCH), not hardcoded
Python-computed values, so the sheet recalculates if raw data changes.
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
import os

BASE = os.path.dirname(__file__)
SRC_CSV = os.path.join(BASE, "..", "data", "processed", "excel_source_q4_2024.csv")
OUT_PATH = os.path.join(BASE, "RetailPulse_Validation_Report.xlsx")

df = pd.read_csv(SRC_CSV)

# ---------------------------------------------------------------
# Styling constants
# ---------------------------------------------------------------
FONT_NAME = "Arial"
HEADER_FILL = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=16, color="2E86AB")
SUBTITLE_FONT = Font(name=FONT_NAME, italic=True, size=10, color="666666")
LABEL_FONT = Font(name=FONT_NAME, bold=True, size=11)
NORMAL_FONT = Font(name=FONT_NAME, size=10)
THIN_BORDER = Border(*[Side(style="thin", color="CCCCCC")] * 4)
KPI_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

wb = openpyxl.Workbook()

# =================================================================
# SHEET 1: Raw Data
# =================================================================
ws_raw = wb.active
ws_raw.title = "Raw Data"

headers = ["Store Name", "Region", "Store Type", "Category", "Month Name", "Month",
           "Revenue", "Cost", "Units Sold", "Transactions"]
ws_raw.append(headers)
for cell in ws_raw[1]:
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center")

for _, row in df.iterrows():
    ws_raw.append([
        row["store_name"], row["region"], row["store_type"], row["category"],
        row["month_name"], int(row["month"]), float(row["revenue"]), float(row["cost"]),
        int(row["units"]), int(row["transactions"])
    ])

# Formatting
for col_idx, col_name in enumerate(headers, 1):
    col_letter = get_column_letter(col_idx)
    ws_raw.column_dimensions[col_letter].width = 20 if col_idx <= 4 else 14
    for r in range(2, len(df) + 2):
        cell = ws_raw[f"{col_letter}{r}"]
        cell.font = NORMAL_FONT
        cell.border = THIN_BORDER
        if col_name in ("Revenue", "Cost"):
            cell.number_format = '#,##0.00'

ws_raw.freeze_panes = "A2"
n_rows = len(df) + 1  # includes header

# =================================================================
# SHEET 2: Regional Summary (SUMIFS validation)
# =================================================================
ws_reg = wb.create_sheet("Regional Summary")
ws_reg["A1"] = "RetailPulse — Regional Performance Summary (Q4 2024)"
ws_reg["A1"].font = TITLE_FONT
ws_reg["A2"] = "Cross-validation of SQL / Power BI regional totals using Excel SUMIFS"
ws_reg["A2"].font = SUBTITLE_FONT

reg_headers = ["Region", "Total Revenue", "Total Cost", "Gross Profit", "Margin %", "Total Units", "Transactions"]
ws_reg.append([])  # row 3 blank
ws_reg.append(reg_headers)  # row 4
for cell in ws_reg[4]:
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center")

regions = sorted(df["region"].unique())
start_row = 5
for i, region in enumerate(regions):
    r = start_row + i
    ws_reg[f"A{r}"] = region
    ws_reg[f"B{r}"] = f"=SUMIFS('Raw Data'!G2:G{n_rows},'Raw Data'!B2:B{n_rows},A{r})"
    ws_reg[f"C{r}"] = f"=SUMIFS('Raw Data'!H2:H{n_rows},'Raw Data'!B2:B{n_rows},A{r})"
    ws_reg[f"D{r}"] = f"=B{r}-C{r}"
    ws_reg[f"E{r}"] = f"=IFERROR(D{r}/B{r},0)"
    ws_reg[f"F{r}"] = f"=SUMIFS('Raw Data'!I2:I{n_rows},'Raw Data'!B2:B{n_rows},A{r})"
    ws_reg[f"G{r}"] = f"=SUMIFS('Raw Data'!J2:J{n_rows},'Raw Data'!B2:B{n_rows},A{r})"
    for col in "ABCDEFG":
        ws_reg[f"{col}{r}"].font = NORMAL_FONT
        ws_reg[f"{col}{r}"].border = THIN_BORDER
    ws_reg[f"B{r}"].number_format = '#,##0.00'
    ws_reg[f"C{r}"].number_format = '#,##0.00'
    ws_reg[f"D{r}"].number_format = '#,##0.00'
    ws_reg[f"E{r}"].number_format = '0.0%'

total_row = start_row + len(regions)
ws_reg[f"A{total_row}"] = "TOTAL"
ws_reg[f"A{total_row}"].font = LABEL_FONT
for col, src_range in [("B", "B"), ("C", "C"), ("D", "D"), ("F", "F"), ("G", "G")]:
    ws_reg[f"{col}{total_row}"] = f"=SUM({col}{start_row}:{col}{total_row-1})"
    ws_reg[f"{col}{total_row}"].font = LABEL_FONT
    ws_reg[f"{col}{total_row}"].number_format = '#,##0.00'
ws_reg[f"E{total_row}"] = f"=IFERROR(D{total_row}/B{total_row},0)"
ws_reg[f"E{total_row}"].font = LABEL_FONT
ws_reg[f"E{total_row}"].number_format = '0.0%'

for col_idx in range(1, 8):
    ws_reg.column_dimensions[get_column_letter(col_idx)].width = 16

# Conditional formatting: highlight regions below 15% margin
ws_reg.conditional_formatting.add(
    f"E{start_row}:E{total_row-1}",
    CellIsRule(operator="lessThan", formula=["0.15"], fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"))
)

# =================================================================
# SHEET 3: Category Profitability
# =================================================================
ws_cat = wb.create_sheet("Category Profitability")
ws_cat["A1"] = "RetailPulse — Category Profitability Analysis (Q4 2024)"
ws_cat["A1"].font = TITLE_FONT
ws_cat["A2"] = "Validates category-level margin figures shown in Power BI Profitability page"
ws_cat["A2"].font = SUBTITLE_FONT

cat_headers = ["Category", "Total Revenue", "Total Cost", "Gross Profit", "Margin %", "Units Sold", "Avg Revenue/Unit"]
ws_cat.append([])
ws_cat.append(cat_headers)
for cell in ws_cat[4]:
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center")

categories = sorted(df["category"].unique())
start_row = 5
for i, cat in enumerate(categories):
    r = start_row + i
    ws_cat[f"A{r}"] = cat
    ws_cat[f"B{r}"] = f"=SUMIFS('Raw Data'!G2:G{n_rows},'Raw Data'!D2:D{n_rows},A{r})"
    ws_cat[f"C{r}"] = f"=SUMIFS('Raw Data'!H2:H{n_rows},'Raw Data'!D2:D{n_rows},A{r})"
    ws_cat[f"D{r}"] = f"=B{r}-C{r}"
    ws_cat[f"E{r}"] = f"=IFERROR(D{r}/B{r},0)"
    ws_cat[f"F{r}"] = f"=SUMIFS('Raw Data'!I2:I{n_rows},'Raw Data'!D2:D{n_rows},A{r})"
    ws_cat[f"G{r}"] = f"=IFERROR(B{r}/F{r},0)"
    for col in "ABCDEFG":
        ws_cat[f"{col}{r}"].font = NORMAL_FONT
        ws_cat[f"{col}{r}"].border = THIN_BORDER
    for col in "BCDG":
        ws_cat[f"{col}{r}"].number_format = '#,##0.00'
    ws_cat[f"E{r}"].number_format = '0.0%'

for col_idx in range(1, 8):
    ws_cat.column_dimensions[get_column_letter(col_idx)].width = 18

# Color scale on margin % to visually flag low-margin categories
cat_total_row = start_row + len(categories) - 1
ws_cat.conditional_formatting.add(
    f"E{start_row}:E{cat_total_row}",
    ColorScaleRule(start_type="min", start_color="FFC7CE", end_type="max", end_color="C6EFCE")
)

wb.save(OUT_PATH)
print(f"Saved base workbook (sheets 1-3) to {OUT_PATH}")
